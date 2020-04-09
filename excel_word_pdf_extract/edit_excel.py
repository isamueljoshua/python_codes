import openpyxl

# create workbook object
wb = openpyxl.Workbook()

# get all sheet names
print(wb.get_sheet_names())

# fetch the sheet by it's name
sheet = wb.get_sheet_by_name('Sheet')
print(sheet)

print(sheet['A1'].value)
print(sheet['A1'].value == None)

# Assigning values to columns
sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('written_excel.xlsx')

# to read from excel file
# openpyxl.load_workbook('written_excel.xlsx')

# Adding sheets to excel file
sheet2 = wb.create_sheet()
print(wb.get_sheet_names())

# change the title of the sheet2 object
sheet2.title = 'My new sheet Name'
print(wb.get_sheet_names())
wb.save('written_excel2.xlsx')

# create sheet as the first one in excel
wb.create_sheet(index=0, title='My Other Sheet')
wb.save('written_excel3.xlsx')
