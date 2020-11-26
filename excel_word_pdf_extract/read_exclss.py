# coding: utf-8

__author__ = ["Samuel Joshua"]

'''
Open a workbook and view with openpyxl
'''
import openpyxl

# to open the excel workbook
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

# to get single sheet from workbook
# the below method is deprecated
# sheet = workbook.get_sheet_by_name('Sheet1')
# use this
sheet = workbook['Sheet1']
print(type(sheet))

# to get all sheet names
# the below is deprecated method, use the next one
# sheets_name_ls = workbook.get_sheet_names()
# use this
sheets_name_ls = workbook.sheetnames
print(sheets_name_ls)

cell = sheet['A1']
print(cell)
print(type(cell.value))
print(cell.value)

print(type(sheet['B1'].value),sheet['B1'].value)
print(type(sheet['C1'].value),sheet['C1'].value)
print(type(str(sheet['C1'].value)),str(sheet['C1'].value))

# rows, columns always start with 1 and not 0
print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

